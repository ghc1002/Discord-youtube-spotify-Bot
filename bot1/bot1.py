
from calendar import month
from datetime import timedelta
import calendar
import discord
import datetime
import openpyxl
import random
import spotipy
import spotipy.util as util
from openpyxl import Workbook
import webbrowser
from re import T, search
from googleapiclient.discovery import build
import os

#path = "wordCount.xlsx"  #used for /compileBot command
       
DEVELOPER_KEY = 'youtube/google dev key'
YOUTUBE_API_SERVICE_NAME = 'youtube'
YOUTUBE_API_VERSION = 'v3'

intents = discord.Intents.default()
intents.message_content = True

client = discord.Client(intents=intents)

url = ""

@client.event
async def on_ready():
    print(f'We have logged in as {client.user}')

@client.event
async def on_message(message):
    if __name__ == '__main__':
        username = "dragonicdeath54"
        token = util.prompt_for_user_token(username,scope = 'streaming user-read-currently-playing',client_id='spotipy client id',client_secret='spotipy client secret',redirect_uri='http://google.com/callback/')
        if token:
           sp = spotipy.Spotify(auth=token)
           
    if message.author == client.user:
        return

    if message.content.startswith('/joinBot') and message.author.voice != None: #have the bot join you in vc
        callingUser = message.author
        VC = callingUser.voice.channel
        await VC.connect()
        
    if message.content.startswith('//pause') and message.author.voice!= None: # pause spotify
        try:
            sp.pause_playback()
            await message.channel.send("Paused")
        except:
            await message.channel.send("spotify offline: please use command //spotify")
        
    if message.content.startswith('//mute') and message.author.voice!= None: #mute spotify
        try:
            sp.volume(0)
        except:
            await message.channel.send("spotify offline: please use command //spotify")
        
    if message.content.startswith('//unmute') and message.author.voice!= None: #unmute spotify
        try:
            sp.volume(100)
        except:
            await message.channel.send("spotify offline: please use command //spotify")
        
    if message.content.startswith('//next') and message.author.voice!= None: #skip to next spotify track
            try:    
                sp.next_track()
                await message.channel.send("Skipped")
            except:
                await message.channel.send("spotify offline: please use command //spotify")
            
    if message.content.startswith('//unpause') and message.author.voice!= None: #unpause spotify
            try:    
                sp.start_playback()
                await message.channel.send("Playing")
            except:
                await message.channel.send("spotify offline: please use command //spotify")
            
    if message.content.startswith('//current') and message.author.voice!= None: #title and artist of currently playing spotify song
        try:
            current = sp.currently_playing().get('item')
            await message.channel.send("Name - " + current.get('name') + "    artist - " + current.get('artists')[0].get('name'))
        except:
            await message.channel.send("spotify offline: please use command //spotify")

    if message.content.startswith('//spotify') and message.author.voice != None: #search for a song on spotify then open it in a new tab
        os.system("taskkill /im msedge.exe /f")
        counter = 1
        trackName = message.content[10:]
        items = sp.search(q='track:' + trackName, type='track', limit=5)
        for track in items.get('tracks').get('items'):
            await message.channel.send(str(counter) + ": Name - " + track.get('name') + "    artist - " + track.get('artists')[0].get('name') + "\n")
            counter = counter + 1
            
        def check(m):
            return m.author == message.author and m.content.isdigit()
        
        msg = await client.wait_for("message", check=check, timeout=60)
        msg = msg.content.strip()
        if msg == '6':
            await message.channel.send("returning")
            return
        vID = items.get('tracks').get('items')[int(msg)-1].get('id')
        await message.channel.send("added to queue")
        
        url = "https://open.spotify.com/track/" + vID
        browser = 'C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe %s'
  
        webbrowser.get(browser)
        webbrowser.open_new(url)
        
    if message.content.startswith('//playlist') and message.author.voice != None and not message.content.startswith('//playlistyt'): #search for a spotify playlist
        os.system("taskkill /im msedge.exe /f")
        counter = 1
        trackName = message.content[11:]
        items = sp.search(q='playlist:' + trackName, type='playlist', limit=5)
        for playlist in items.get('playlists').get('items'):
            await message.channel.send(str(counter) + ": Name - " + playlist.get('name') + "    artist - " + playlist.get('owner').get('display_name') + "\n")
            counter = counter + 1
            
        def check(m):
            return m.author == message.author and m.content.isdigit()
        
        msg = await client.wait_for("message", check=check, timeout=60)
        msg = msg.content.strip()
        if msg == '6':
            await message.channel.send("returning")
            return
        vID = items.get('playlists').get('items')[int(msg)-1].get('id')
        await message.channel.send("added to queue")
        
        url = "https://open.spotify.com/playlist/" + vID
        browser = 'C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe %s'
  
        webbrowser.get(browser)
        webbrowser.open_new(url)

    if message.content.startswith('//add') and message.author.voice != None: #search for a song based on title then add it to spotify queue
        counter = 1
        trackName = message.content[6:]
        items = sp.search(q='track:' + trackName, type='track', limit=5)
        for track in items.get('tracks').get('items'):
            await message.channel.send(str(counter) + ": Name - " + track.get('name') + "    artist - " + track.get('artists')[0].get('name') + "\n")
            counter = counter + 1
            
        def check(m):
            return m.author == message.author and m.content.isdigit()
        
        msg = await client.wait_for("message", check=check, timeout=60)
        msg = msg.content.strip()
        if msg == '6':
            await message.channel.send("returning")
            return
        uri = items.get('tracks').get('items')[int(msg)-1].get('uri')
        await message.channel.send("added to queue")
        try:
            sp.add_to_queue(uri)
            sp.volume(100)
        except:
            await message.channel.send("spotify offline: please use command //spotify")
        
    if message.content.startswith('//play') and message.author.voice != None and not message.content.startswith('//play http') and not message.content.startswith('//playlist'): #search for a youtube video than plays it
        def openYT(vID):
          url = "https://www.youtube.com/watch?v=" + vID
          browser = 'C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe %s'
  
          webbrowser.get(browser)
          webbrowser.open_new(url)

        os.system("taskkill /im msedge.exe /f")
        youtube = build(YOUTUBE_API_SERVICE_NAME, YOUTUBE_API_VERSION,
        developerKey=DEVELOPER_KEY)
        
        title = message.content[7:]

        # Call the search.list method to retrieve results matching the specified
        # query term.
        search_response = youtube.search().list(
          q=title,
          part='snippet',
          maxResults=5,
          type = 'video'
        ).execute()
        items = search_response['items']
        selection = 0
        counter = 1
        for item in items:
          vid = youtube.videos().list(
            id = item['id']['videoId'],
            part = 'contentDetails',
            maxResults = 1
          ).execute()
          duration = vid['items'][0]['contentDetails']['duration']
          duration = duration[2:]
          if(duration.find('H') != -1):
              duration = duration.replace('H', 'H ')
          if(duration.find('M') != -1):
            duration = duration.replace('M', ':')
            duration = duration.replace('S', '')
          if(duration.find('S')):
            duration = duration.replace('S', ' Seconds')
          await message.channel.send(str(counter) + ': ' + item['snippet']['title'] + "  (" + duration + ")\nAuthor: " + item['snippet']['channelTitle'])
          counter = counter + 1
        
        def check(m):
            return m.author == message.author and m.content.isdigit()
        
        selection = await client.wait_for("message", check=check, timeout=60)
        selection = selection.content.strip()
        if selection == '0':
            await message.channel.send("returning")
            return
        
        choice = items[int(selection) - 1]
        ID = choice['id']
        vID = ID['videoId']
        await message.channel.send("playing...")
        openYT(vID)
        
    if message.content.startswith('//ytplaylist') and message.author.voice != None: #searches for a youtube playlist
        def openYT(vID):
          url = "https://www.youtube.com/playlist?list=" + vID
          browser = 'C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe %s'
  
          webbrowser.get(browser)
          webbrowser.open_new(url)

        os.system("taskkill /im msedge.exe /f")
        youtube = build(YOUTUBE_API_SERVICE_NAME, YOUTUBE_API_VERSION,
        developerKey=DEVELOPER_KEY)
        
        title = message.content[13:]

        # Call the search.list method to retrieve results matching the specified
        # query term.
        search_response = youtube.search().list(
          q=title,
          part='snippet',
          maxResults=5,
          type = 'playlist'
        ).execute()
        items = search_response['items']
        selection = 0
        counter = 1
        for item in items:
          await message.channel.send(str(counter) + ': ' + item['snippet']['title'] + "\nAuthor: " + item['snippet']['channelTitle'])
          counter = counter + 1
        
        def check(m):
            return m.author == message.author and m.content.isdigit()
        
        selection = await client.wait_for("message", check=check, timeout=60)
        selection = selection.content.strip()
        if selection == '0':
            await message.channel.send("returning")
            return
        
        choice = items[int(selection) - 1]
        ID = choice['id']
        vID = ID['playlistId']
        await message.channel.send("playing...")
        openYT(vID)
        
    if message.content.startswith('//play http') and message.author.voice != None and not message.content.startswith('//playlist'): #plays a link
        os.system("taskkill /im msedge.exe /f")
        link = message.content[7:]
        url = link
        browser = 'C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe %s'
  
        webbrowser.get(browser)
        webbrowser.open_new(url)
        
    if message.content.startswith('//ytstop') and message.author.voice != None: #close browser
        os.system("taskkill /im msedge.exe /f")
        
    if message.content.startswith('/quote'): #sends a random message that has been sent at some point the channel
        channel = message.channel
        creationDate = channel.created_at
        now = datetime.datetime.now()
        year = random.randrange(creationDate.year, now.year+1)
        month = random.randrange(1, 12)
        if year == now.year:
            month = random.randrange(1, now.month)
        elif year == creationDate.year:
            month = random.randrange(creationDate.month, 12)
        dayNum = calendar.monthrange(year, month)
        day = random.randrange(0, dayNum[1]-1)
        if year == now.year and month == now.month:
            day = random.randrange(0, now.day)
        elif year == creationDate.year and month == creationDate.month:
            day = random.randrange(creationDate.day, dayNum[1])
        hour = random.randrange(0, 24)
        if year == now.year and month == now.month and now.day == day:
            day = random.randrange(0, now.hour+1)
        elif year == creationDate.year and month == creationDate.month and day == creationDate.day:
            day = random.randrange(creationDate.hour, 24)
        minute = random.randrange(0, 60)
        if year == now.year and month == now.month and now.day == day and hour == now.hour:
            day = random.randrange(0, now.minute+1)
        elif year == creationDate.year and month == creationDate.month and day == creationDate.day and hour == creationDate.hour:
            day = random.randrange(creationDate.minute, 60)
        second = random.randrange(0, 60)
        date = datetime.datetime(year, month, day, hour=hour, minute=minute, second=second)
        async for quote in channel.history(limit=1, around = date):
            date = quote.created_at - timedelta(hours=4)
            url = ""
            if quote.attachments != []:
                for attach in quote.attachments:
                    url = attach.url
                    await message.channel.send(url)
            minute = str(date.minute)
            if int(minute) < 10:
                minute = "0" + str(date.minute)
            quote = quote.content + "\n\nsent by: " + quote.author.name + "\nsent on the " + str(date.day) + " of " + date.strftime("%B") + " " + str(date.year) + " at " + str(date.hour) + ":" + minute
            await message.channel.send(content = quote, allowed_mentions = discord.AllowedMentions.none())

    '''if message.content.startswith('/compileBot'): #command to find every word ever said in the channel count how many times each word has appeared and add them all to an ms excel sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        row = 1
        channel = message.channel
        wordPile = {}
        async for quote in channel.history(limit = None):
            sentence = quote.content.split()
            print(sentence)
            for word in sentence:
                if not word.isalpha():
                    continue
                
                word = word.lower()
                print(word)
                if wordPile.get(word) != None:
                    wordPile.update({word: wordPile.get(word)+1})
                else:
                    wordPile[word] = 1
            
        await channel.send("wording done")
        for word in wordPile.keys():
            col1 = sheet.cell(row = row, column = 1)
            col2 = sheet.cell(row = row, column = 2)
            col1.value = word
            col2.value = wordPile.get(word)
            row = row+1
         
        await channel.send("all done")
        wb.save("wordCount.xlsx") '''

client.run('discord dev key')